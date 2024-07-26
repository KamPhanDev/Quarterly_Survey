import openpyxl as xl
import Config

wb = xl.load_workbook(Config.fileName)
print(wb.sheetnames)
ws = wb[Config.responseData]

maxRow, maxCol = ws.max_row +1, ws.max_column + 1

print(f'Row:{maxRow}, Col:{maxCol}')

feedbackList = []
improvementList = []
feelingList = []


for i in range(2, ws.max_row):
    if ws.cell(i, 26).value != None:
        words = ws.cell(i, 26).value
        try:
            words = words.split(', ')
        except:
            continue

        feelingList += words
        words.clear
        
feelingSet = set(feelingList)

with open('feelingKeyword.txt', 'w') as feeling:
    feeling.write(str(feelingSet))

for item in feelingSet:
    print(item)

print('=' * 80)

for i in range(2, ws.max_row+1):
    if ws.cell(i, 25).value != None:
        words = ws.cell(i, 25).value
        try:
            words = words.split(', ')
        except:
            continue
        
        improvementList += words
        words.clear
        
improvementSet = set(improvementList)

with open('improvementKeyword.txt', 'w') as improvement:
    improvement.write(str(improvementSet))

for item in improvementSet:
    print(item)

print('=' * 80)

for i in range(2, ws.max_row+1):
    if ws.cell(i, 24).value != None:
        words = ws.cell(i, 24).value
        try:
            words = words.split(', ')
        except:
            continue

        feedbackList += words
        words.clear
        
feedbackSet = set(feedbackList)

with open('feedbackKeyword.txt','w') as feedback:
    feedback.write(str(feedbackSet))

for item in feedbackSet:
    print(item)
