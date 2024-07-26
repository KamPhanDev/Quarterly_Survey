#! Python3

import re
import openpyxl as xl
# import pandas as pd
import nltk
from nltk.corpus import stopwords
from nltk.sentiment import SentimentIntensityAnalyzer
from icecream import ic
import Config
nltk.download('vader_lexicon')
ic.enable()

sia = SentimentIntensityAnalyzer()

def getTextData(fileName):
    with open(fileName, 'r', encoding=None) as txt:
        ref = txt.readline().lower()
        return ref.split(',')

refFeedback = getTextData('feedbackKeyword.txt')
ic(refFeedback)

refImprovement = getTextData('improvementKeyword.txt')
ic(refImprovement)

refeeling = getTextData('feelingKeyword.txt')
ic(refeeling)

wb = xl.load_workbook(str(Config.fileName))
ic(wb)

try:
    ws = wb[Config.responseData]
except:
    ws = wb['Sheet1']

maxCol = ws.max_column + 1
maxRow = ws.max_row +1

ic(stopwords.words("english"))

for i in range(7858, maxRow):
    if ws.cell(i, Config.feedbackCol).value is not None:
        ic(ws.cell(i, Config.feedbackCol).value)
        textStr = re.sub(pattern="[^\w\s]", repl="", string=ws.cell(i, Config.feedbackCol).value)

        tokenText = textStr.split()
        stopText = [word for word in tokenText if word.lower() not in stopwords.words('english')]
        text = [word for word in stopText if word.lower() in refFeedback]

        ic(text)
        ws.cell(i, Config.feedbackKeyCol).value = ','.join(text)
        sentScore = sia.polarity_scores(ws.cell(i, Config.feedbackCol).value)
        ws.cell(i, Config.feedbackSent).value = sentScore['compound']

    if ws.cell(i, Config.improvementCol).value is not None:
        ic(ws.cell(i, Config.improvementCol).value)
        textStr = re.sub(pattern="[^\w\s]", repl="", string=ws.cell(i, Config.improvementCol).value)

        tokenText = textStr.split()
        stopText = [word for word in tokenText if word.lower() not in stopwords.words('english')]
        text = [word for word in stopText if word.lower() in refImprovement]

        ic(text)
        ws.cell(i, Config.improvementKeyCol).value = ','.join(text)
        sentScore = sia.polarity_scores(ws.cell(i, Config.improvementCol).value)
        ws.cell(i, Config.improvementSent).value = sentScore['compound']

    if ws.cell(i, Config.feelingCol).value is not None:
        ic(ws.cell(i, Config.feelingCol).value)
        textStr = re.sub(pattern="[^\w\s]", repl="", string=ws.cell(i, Config.feelingCol).value)

        tokenText = textStr.split()
        stopText = [word for word in tokenText if word.lower() not in stopwords.words('english')]
        text = [word for word in stopText if word.lower() in refeeling]

        ic(text)
        ws.cell(i, Config.feelingKeyCol).value = ','.join(text)
        sentScore = sia.polarity_scores(ws.cell(i, Config.feelingCol).value)
        ws.cell(i, Config.feelingSent).value = sentScore['compound']

wb.save(Config.fileName)
wb.close